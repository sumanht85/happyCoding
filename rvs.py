import streamlit as st
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime

# Load Excel file
def load_excel(file_name):
    book = load_workbook(file_name)
    return book

# Save data to Excel
def save_to_excel(book, file_name):
    book.save(file_name)

# Register a new user
def register_user(book, username, password):
    users_sheet = book['Users']
    users_sheet.append([username, password])
    save_to_excel(book, 'expenses.xlsx')

# Authenticate user
def authenticate_user(book, username, password):
    users_sheet = book['Users']
    for row in users_sheet.iter_rows(values_only=True):
        if row[0] == username and row[1] == password:
            return True
    return False

# Add expense/investment
def add_record(book, username, date, record_type, description, amount):
    expenses_sheet = book['Expenses']
    expenses_sheet.append([username, date, record_type, description, amount])
    save_to_excel(book, 'expenses.xlsx')

# Get user expenses
def get_user_records(book, username):
    expenses_sheet = book['Expenses']
    data = []
    for row in expenses_sheet.iter_rows(values_only=True):
        if row[0] == username:
            data.append(row)
    return pd.DataFrame(data, columns=['Username', 'Date', 'Type', 'Description', 'Amount'])

# Calculate balance
def calculate_balance(df):
    income = df[df['Type'] == 'Investment']['Amount'].sum()
    expenses = df[df['Type'] == 'Expense']['Amount'].sum()
    balance = income - expenses
    return income, expenses, balance

# Load the Excel book
book = load_excel('expenses.xlsx')

# Streamlit UI
st.title("House Building Expense Tracker")

if 'logged_in' not in st.session_state:
    st.session_state['logged_in'] = False
if 'current_page' not in st.session_state:
    st.session_state['current_page'] = "Login"

def switch_page(page):
    st.session_state['current_page'] = page

menu = ["Login", "Register", "Add Record", "View Reports"]
choice = st.sidebar.selectbox("Menu", menu)

if choice == "Register":
    st.subheader("Create a New Account")
    username = st.text_input("Username")
    password = st.text_input("Password", type='password')
    if st.button("Register"):
        register_user(book, username, password)
        st.success("You have successfully registered")
        switch_page("Login")
elif choice == "Login":
    st.subheader("Login to Your Account")
    username = st.text_input("Username")
    password = st.text_input("Password", type='password')
    if st.button("Login"):
        if authenticate_user(book, username, password):
            st.success("Logged in successfully")
            st.session_state['logged_in'] = True
            st.session_state['username'] = username
            switch_page("View Reports")
        else:
            st.error("Invalid Username/Password")
elif choice == "Add Record" and st.session_state.get('logged_in'):
    st.subheader("Add a New Record")
    date = st.date_input("Date", datetime.now())
    record_type = st.selectbox("Type", ["Investment", "Expense"])
    description = st.text_input("Description")
    amount = st.number_input("Amount", min_value=0.0)
    if st.button("Add"):
        add_record(book, st.session_state['username'], date, record_type, description, amount)
        st.success("Record added successfully")
        switch_page("View Reports")
elif st.session_state['logged_in'] and st.session_state['current_page'] == "View Reports":
    st.subheader("Your Reports")
    df = get_user_records(book, st.session_state['username'])
    
    # Show date-wise report
    st.write("### Date-wise Report")
    datewise_report = df.groupby('Date')['Amount'].sum().reset_index()
    st.dataframe(datewise_report)
    
    # Show month-wise report
    st.write("### Month-wise Report")
    df['Month'] = pd.to_datetime(df['Date']).dt.to_period('M')
    monthwise_report = df.groupby('Month')['Amount'].sum().reset_index()
    st.dataframe(monthwise_report)
    
    # Show balance
    income, expenses, balance = calculate_balance(df)
    st.write(f"### Summary")
    st.write(f"Total Investments: {income}")
    st.write(f"Total Expenses: {expenses}")
    st.write(f"Balance: {balance}")
else:
    st.warning("Please login to access this feature.")
